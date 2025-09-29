import pandas as pd
import time
from datetime import datetime
from lengthy_imports import *

from dateutil.relativedelta import relativedelta
import numpy as np
from functools import reduce

import seaborn as sns
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.patches as mpatches
from matplotlib.ticker import FixedLocator

import ast
import openpyxl as pyxl
from openpyxl.styles import DEFAULT_FONT
import os
import copy

import gui_functions as gui_f

from ipydex import IPS, activate_ips_on_exception

from typing import Union, Optional
idx = pd.IndexSlice

def load_checklist() -> pd.DataFrame:
    return pd.read_csv(f'{mainpath}/databases/checklist.csv',
                       sep=';', index_col=[0, 1, 2])
    
def load_remarks() -> pd.DataFrame:
    # return pd.read_hdf(f'{mainpath}/databases/remarks.h5', key='remarks')
    return pd.read_csv(f'{mainpath}/databases/remarks.csv',
                       sep=';', index_col=[0, 1, 2, 3, 4, 5],
                       parse_dates=['create_time'])


def load_parts() -> pd.DataFrame:
    return pd.read_csv(f'{mainpath}/databases/components.csv',
                    sep=';', index_col=[0, 1, 2, 3])


def load_temperatures() -> pd.DataFrame:
    return pd.read_csv(f'{mainpath}/databases/temperatures.csv',
                       sep=';', index_col=[0, 1, 2, 3, 4])


def load_turbines() -> pd.DataFrame:
    return pd.read_csv(f'{mainpath}/databases/turbines.csv',
                       sep=';', index_col=[0, 1])


def load_inspections() -> pd.DataFrame:
    return pd.read_csv(f'{mainpath}/databases/inspections.csv',
                       sep=';', index_col=[0, 1, 2, 3])


def save_db(db: pd.DataFrame|pd.Series, name:str):
    path = f'{mainpath}/databases/{name}.csv'
    db.sort_index().to_csv(path, sep=';')

def get_order(what):
    '''return ordered list/dict
    what[str], 'chapters', 'components', 'temperatures'
    '''
    if what not in ['chapters', 'components', 'temperatures']:
        raise ValueError("cannot understand what to parse. must be one of "
                         "'chapters', 'components', 'temperatures'")
    path = f'{mainpath}/databases/order_{what}.txt'
    with open(path, 'r', encoding='utf-8') as f:
        data = f.read()
    return ast.literal_eval(data)

def order_by_position(remarks):
    if remarks is None:
        return remarks
    if remarks.empty:
        return remarks
    order_numbers = remarks['position'].dropna()
    if order_numbers.empty:
        return remarks
    position_counts = order_numbers.groupby('address', observed=True).value_counts()
    if position_counts.gt(1).any():
        raise ValueError(f'manually assigned positions overlap: {order_numbers.to_list()} '
                         f'(example index: {order_numbers.index[0]})')
    
    groups = get_chapters_from_remarks(remarks)
    ordered = []
    for group in groups:
        i = 0
        group_rems = get_remarks_of_chapter(remarks, group)
        order_numbers = group_rems['position'].dropna()
        if order_numbers.empty:
            ordered.append(group_rems)
            continue
    
        manual = group_rems[group_rems.position.notna()].sort_values(by='position')
        unpositioned = group_rems[group_rems.position.isna()]

        for position in range(1, int(max([len(group_rems)+1,
                                          manual[['position']].max().values[0]+1]))):
            if position in manual.position.values:
                ordered.append(manual.iloc[[i]])
                i += 1
            elif not unpositioned.empty:
                try:
                    ordered.append(unpositioned.iloc[[0]])
                    unpositioned = unpositioned.iloc[1:]
                except IndexError: continue

    ordered_df = pd.concat(ordered)
    ordered_df.index.names = remarks.index.names

    if not ordered_df.eq(remarks).fulltext.all():
        raise ValueError(f'DataFrame mismatch. Ordered remarks have length {len(ordered)}, '
                         f'original DF has length {len(remarks)}. '
                        'Somethings wrong, perhaps negative or duplicate order numbers?')
    return ordered_df
    

def rempos2str(rempos):
    if rempos is None or rempos == '' or pd.isna(rempos):
        return ''
    if isinstance(rempos, int) or isinstance(rempos, float):
        return int(rempos)
    return f'{int(float(rempos.strip()))}'




def check_if_checklist_entry_exists(address: str,
                          title: Optional[str]=None) -> bool:
    '''return True if given address is in the current checklist
    checklist is reloaded from csv
    address (str): section's or remarkstitle's path
        interpreted as remarktitle's path if title is None
    title (str): title of remark, if not contained in address
    '''
    checklist = load_checklist()
    if title is None:
        # last str after db_split_char (default '|') is interpreted as remark title
        sectionaddress = address[:address.rfind(db_split_char)]
        remarktitle = address.split(db_split_char)[-1]
    else:
        sectionaddress = address
        remarktitle = title
    
    try:
        checklist.loc[sectionaddress, remarktitle]
        return True
    except KeyError:
        return False

def multiindex2dict(p: Union[pd.MultiIndex, dict]) -> dict:
    """
    Converts a pandas Multiindex to a nested dict
    :parm p: As this is a recursive function, initially p is a pd.MultiIndex, but after the first iteration it takes
    the internal_dict value, so it becomes to a dictionary
    """
    internal_dict = {}
    end = False
    for x in p:
        # Since multi-indexes have a descending hierarchical structure, it is convenient to start from the last
        # element of each tuple. That is, we start by generating the lower level to the upper one. See the example
        if isinstance(p, pd.MultiIndex):
            # This checks if the tuple x without the last element has len = 1. If so, the unique value of the
            # remaining tuple works as key in the new dict, otherwise the remaining tuple is used. Only for 2 levels
            # pd.MultiIndex
            if len(x[:-1]) == 1:
                t = x[:-1][0]
                end = True
            else:
                t = x[:-1]
            if t not in internal_dict:
                internal_dict[t] = [x[-1]]
            else:
                internal_dict[t].append(x[-1])
        elif isinstance(x, tuple):
            # This checks if the tuple x without the last element has len = 1. If so, the unique value of the
            # remaining tuple works as key in the new dict, otherwise the remaining tuple is used
            if len(x[:-1]) == 1:
                t = x[:-1][0]
                end = True
            else:
                t = x[:-1]
            if t not in internal_dict:
                internal_dict[t] = {x[-1]: p[x]}
            else:
                internal_dict[t][x[-1]] = p[x]
    
    # Uncomment this line to know how the dictionary is generated starting from the lowest level
    # print(internal_dict)
    if end:
        return internal_dict
    return multiindex2dict(internal_dict)

def multi_level_drop(df, level_names, values):
    """
    Drop rows from a MultiIndex DataFrame based on conditions for multiple levels.
    
    Parameters:
        df (pd.DataFrame): The MultiIndex DataFrame.
        level_names (list): List of level names in the MultiIndex to check.
        values (list): List of corresponding values for the levels. Must match `level_names`.
        
    Returns:
        pd.DataFrame: The filtered DataFrame.
    """
    if len(level_names) != len(values):
        raise ValueError("level_names and values must have the same length")
    
    # Build the mask dynamically
    mask = True
    for level, value in zip(level_names, values):
        mask &= (df.index.get_level_values(level).astype(str) == str(value))
    
    # Negate the mask to keep rows that do not match the condition
    return df.loc[~mask]

def dict2addresses(d):
    l = []
    for key in d.keys():
        l.extend([f'{key}{db_split_char}{i}' for i in d[key]])
    return l

def ser2addresses(ser):
    return (ser.reset_index().address + '|' + ser.reset_index().titles).to_list()


def get_turbine_oem_id_dict():
    db = load_turbines()
    return multiindex2dict(db.set_index('model', append=True)
                           .droplevel('sn')
                           .groupby(['oem', 'model'])
                           .head(1)
                           .index)



def get_component_model_dict_for_turbine_type(oem: str,
                                              model: str,
                                              component_name: str,
                                              ):
    '''returns a dict that contains all component oems and models
    for given turbine
    '''
    parts = load_parts()
            
    supermodel = gui_f.get_supermodel_from_model(model)
    supermodel_in_index = (parts.index
                           .get_level_values('model')
                           .str
                           .startswith(supermodel))
    parts = parts.iloc[supermodel_in_index]

    try:
        known_models_as_ind = (parts
                               .loc[idx[oem, :, component_name]]
                               .reset_index()
                               .set_index(['component_oem', 'component_model'])
                               .index
                               .unique())
        # remove everything after the first ',' in component_model --> eg. blade set number disappears
        if component_name in ['Rotorblätter', ]:
            known_models_as_ind = (pd.MultiIndex
                                .from_arrays((
                                    (known_models_as_ind
                                            .get_level_values('component_oem')),
                                        (known_models_as_ind
                                            .get_level_values('component_model')
                                            .str
                                            .split(', ')
                                            .str[0])),
                                        names=['component_oem', 'component_model'])
                                .unique())
    except KeyError: # occurs when combination of oem, model and component_name don't exist
        return {}

    return multiindex2dict(known_models_as_ind)

def get_sample_sn_from_part_model(
        part_oem: str, part_model: str,
        wea_oem: Optional[str]=None, wea_model: Optional[str]=None,
        wea_sn: Optional[str]=None,
    ) -> str:
    '''gives a sample serial number for a given part type'''
    parts = load_parts()
    same_wea_model = False

    occurrences_of_part = (parts
                           .where(((parts.component_oem==part_oem) & 
                                   (parts.component_model==part_model)))
                           .dropna(how='all')
                           .sn)
    
    if wea_model is not None and wea_oem is not None:
        try:
            occurrences_of_part = occurrences_of_part.loc[wea_oem, wea_model]
            same_wea_model = True
        except KeyError:
            pass # just keep the old occurences_of_part
    # return sn as prefill for wea with closest wea_id
    if wea_sn is not None and same_wea_model:
        try:
            available_sn = (occurrences_of_part
                            .index
                            .get_level_values('turbine_id')
                            .astype(int))
            wea_sn = int(wea_sn)
            closest_id = available_sn[np.abs(available_sn-wea_sn).argmin()]
            return (occurrences_of_part
                    .loc[idx[:, f'{closest_id}']]
                    .sort_index(level='insp_year', ascending=False)
                    .iloc[0])
        except ValueError:
            pass
        
    try:
        sn = occurrences_of_part.iloc[0]
    except IndexError:
        sn = ''
    if pd.isna(sn):
        sn = ''
    elif type(sn) == str:
        if sn.lower() == 'nan' or sn.lower() == 'none':
            sn = ''    
    return sn

def turbine_in_db(db: pd.DataFrame | pd.Series, oem: str, turb_id: str) -> bool:
    try:
        db.xs((oem, turb_id), level=['oem', 'turbine_id'])
        return True
    except KeyError:
        return False
    
def get_turbine_parts(oem, turb_id, parts_db=None, sorted=True) -> dict:
    '''returns a dict containig the most recent entries in the components_db 
    for a given turbine's parts. Returns most recent value for every part
    that has ever been in the DB. So if a part was not mentioned in 2024, but was
    present in the DB in 2022, this part's info is NOT omitted, but instead its
    info from 2022 is put into the returned dict.
    
    returned dict has shape {partname1: [oem, model, sn], partname2: ...}
    '''
    if parts_db is None:
        parts_db = load_parts()
    try:
        db_slice = parts_db.xs((oem, turb_id), level=['oem', 'turbine_id']).astype(str)
    except KeyError:
        return {}
    parts_dict = (db_slice
                  .droplevel('model')
                  .T
                  .to_dict(orient='list'))
    if not sorted: return parts_dict
    return sort_dict(parts_dict, get_order('components'))

def get_sec_title_from_address(address):
    '''splits an address into section and title, return (None, None) if None is given'''
    if not address:
        return None, None
    split = address.rfind(db_split_char)
    title = address[split+1:]
    section = address[:split]
    return section, title

def get_empty_remarks_df():
    return pd.DataFrame(index=pd.MultiIndex.from_tuples([],
                        names=['oem', 'turbine_id', 'insp_year',
                               'inspection_type', 'address', 'create_time']),
                        columns=['fulltext', 'flag', 'image_names', 'author', 'position'])


def get_chapters_from_remarks(remarks: pd.DataFrame, unique: bool=True,
                              as_list=True):
    if remarks is None:
        remarks = get_empty_remarks_df()
    ind = (remarks.index.get_level_values('address')
           .map(lambda x: x[:x.rfind(db_split_char)]))
    if unique: ind = ind.unique()
    if as_list: ind = ind.to_list()
    return ind

def get_chapters_from_checklist(checklist=None) -> list:
    if checklist is None:
        checklist = load_checklist()
    chapters = checklist.index.unique('chapter').to_list()
    return chapters

def get_chapters_of_subpath(list_with_overlap=None,
                            parent_path=None, standalone_name=standalone_chapter_str):
    '''gives a list of subchapters for given parent chapters. Respects the levels as
    separated by the value of db_split_char (default '>')
    parent_path: str or None, only consinder sections that are children to the parent chapter
        if None: get highest level chapters
    standalone_name: name of subhchapter if the checklist entry is directly inside the parent chapter,
        not a hcild section
    returns list of str, list of chapters within a parent chapter
    '''
    if list_with_overlap is None:
        return None
    all_chapters = pd.Series(list_with_overlap).dropna()

    if not parent_path:
        levelchaps = list(all_chapters.str.split(db_split_char).str[0].unique())
        return levelchaps
    
    relevant_chapters = all_chapters.mask(all_chapters.str.find(parent_path)!=0).dropna()
    levelchaps = (relevant_chapters
                  .str.split(parent_path, regex=False).str[1]
                  .str.split(db_split_char).str[1]
                  .replace({pd.NA: standalone_name})
                  .unique())
    if not pd.Series(levelchaps).any():
        return None
    return list(levelchaps)

def db_blwl2list(blwl):
    if blwl is None:
        return []
    if pd.isna(blwl):
        return []
    if len(blwl) == 0:
        return []
    blwl_raw = blwl.split(',')
    return [el.strip() for el in blwl_raw if el.strip()]
    

def checklist_remove_blacklisted(wea, checklist: pd.DataFrame) -> pd.DataFrame:
        '''remove entries from checklist that are blacklisted 
        based on the current turbine type'''
        oem = wea.get('oem')
        model = wea.get('model')
        tower_type = wea.get('tower_type')

        if checklist.blacklist.isnull().all():
            return checklist
        
        isblacklisted = pd.Series([False]*len(checklist), index=checklist.index)

        if oem:
            isblacklisted |= (checklist
                              .blacklist
                              .str.contains(f'{oem},')
                              .fillna(False)
                              .astype(bool)) # comma explanation see checklist_get_whitelisted
        if model:
            isblacklisted |= (checklist
                              .blacklist
                              .str.contains(model)
                              .fillna(False)
                              .astype(bool))
        if tower_type:
            isblacklisted |= reduce(lambda x, y: x | y, [(checklist
                                                    .blacklist
                                                    .str.contains(tower_unit)
                                                    .fillna(False)
                                                    .astype(bool))\
                         for tower_unit in tower_type.split(db_split_char)])


        if isblacklisted.any():
            checklist = checklist.mask(isblacklisted).dropna(subset='fulltext')
        return checklist

def checklist_get_whitelisted(wea, checklist: pd.DataFrame) -> pd.DataFrame:
    tower_type = wea.get('tower_type')
    model = wea.get('model')
    oem = wea.get('oem')

    wl_empty = (checklist
                .whitelist
                .isnull())
    wanted = wl_empty
    if wl_empty.all(): return checklist

    # prepare conditions
    if tower_type:
        wanted |= reduce(lambda x, y: x | y, [(checklist
                                               .whitelist
                                               .str.contains(tower_unit)
                                               .fillna(False)
                                               .astype(bool))\
                         for tower_unit in tower_type.split(db_split_char)])
    if model:
        wanted |= (checklist
                   .whitelist
                   .str.contains(model)
                   .fillna(False)
                   .astype(bool))   # bool necessary for | operation in the end
    if oem:
        wanted |= (checklist
                   .whitelist
                   .str.contains(f'{oem},') # comma to not include other whitelisted models of same oem (do not onclude Enercon E70 if wl Enercon is asked)
                   .fillna(False)
                   .astype(bool))
        
    return checklist.where(wanted).dropna(subset='fulltext')

def checklist_get_selected_chapters(report, checklist) -> pd.DataFrame:
    chapters = report.chapters
    return checklist.loc[chapters]

def get_report_checklist(report) -> pd.DataFrame:
    '''
    - only shows appropriate checklist entries according to blacklist and whitelists
    '''
    wea = report.parent_wea
    checklist = load_checklist()
    try:
        checklist_wl = checklist_get_whitelisted(wea, checklist)
    except Exception as e:
        IPS()
        raise e
    cl = checklist_remove_blacklisted(wea, checklist_wl)
    cl = cl.loc[~cl.index.duplicated()]
    return cl

def filter_from_turb_db(l: list, what: str) -> list:
    '''return 'what' from l, eg. oems from the given list. Uses turbine_db to
    find models and oems
    what: string, (oem, model, tower_type)
    l: list of strings from which to filter'''
    if what not in ('oem', 'model', 'tower_type'):
        raise AttributeError(f'what must be oem, model or tower_type, but is {what}')
    if not isinstance(l, list):
        raise ValueError(f'l must be list, but is {l}')
    wanted = get_all(what)
    return [i for i in l if i in wanted]

def sort_remarks_using_address(df: pd.DataFrame) -> pd.DataFrame:
    '''sorts a dataframe with 'address' as one index level according to
    order_chapters.txt
    checks for missing entries after sorting'''
    df_og = copy.deepcopy(df)
    indexnames_before = list(df.index.names)
    addresses_before = df.index.get_level_values('address')
    new_order = ['address']
    new_order.extend([level for level in indexnames_before if level != 'address'])
    df = df.reorder_levels(new_order)
    chap_order = get_order('chapters')
    order = []
    for chap in chap_order.keys():
        titles = chap_order[chap]
        for title in titles:
            order.append(f'{chap}|{title}')
    df = (df
          .reindex(order, level='address')
          .dropna(how='all')
          .reorder_levels(indexnames_before))

    # catch faults that occurred while sorting
    addresses_after = df.index.get_level_values('address')
    for original_address in addresses_before:
        if original_address not in addresses_after:
            raise IndexError(f'Remark {original_address} has been removed. Are both Section and Title exactly present in order_chapters.txt?')

    if len(df_og) != len(df):
        raise IndexError(f'No address has been dropped while sorting, still at least one remark is missing from the overview... Abort. Enter console and find the bug, or contact Tade')
    for ind in df_og.index:
        og = df_og.loc[ind]
        new = df.loc[ind]
        if not (og.dropna() == new.dropna()).all():
            raise ValueError(f'Something\'s gone mixed up at index {ind}. Please check manually.')
    return df

def get_project_remarks(project,
                        allowed_flags: Optional[list]=['PPP', 'PP', 'P',
                                                       'I', 'V', 'E'],
                        drop_cols: Optional[list]=[]) -> pd.DataFrame:
    '''to ex-/include certain types of remarks, set allowed_flags,
    drop certain columns'''
    wea_ids = project.windfarm.get_setup_wea_ids()
    year_id = project.get('year_id')

    rem_db = load_remarks()

    single_weas_rems = []
    for wea_id in wea_ids:
        wea = project.windfarm.get_wea(wea_id)
        kind = wea.report.inspection.kind
        try:
            locs = rem_db.index.get_locs(idx[wea.oem, wea_id, year_id, kind])
            single_wea_rems = rem_db.iloc[locs]
            single_weas_rems.append(single_wea_rems)
        except KeyError:
            continue
    
    all_park_rems = pd.DataFrame()
    if single_weas_rems:
        all_park_rems = pd.concat(single_weas_rems, axis=0)
    if drop_cols: all_park_rems = all_park_rems.drop(drop_cols, axis=1)
    if not allowed_flags: return all_park_rems

    # filter for actual remarks only (drop conclusion etc)
    allowed_entries = reduce(lambda x, y: x|y,
                             [all_park_rems.flag == flag for flag in allowed_flags])
    all_park_rems = all_park_rems[allowed_entries]
    return all_park_rems

def get_remarks_from_ids(oem, ids, remarks=None):
    '''get DataFrame of all remarks with given oem and ids.
    remarks: remarks DataFrame as returned by load_remarks'''
    r = remarks if remarks is not None else load_remarks()
    if not oem in r.index.unique('oem'): return get_empty_remarks_df()

    r_oem = r.loc[oem]
    present_ids = r_oem.index.unique('turbine_id').intersection(ids)

    if present_ids.empty: return get_empty_remarks_df()

    return r.loc[idx[oem, present_ids], ]


def filter_specific_report(remarks, wea) -> pd.DataFrame:
    '''takes a DataFrame of remarks and filters for the given wea's report.
    removes unneccessary index levels. The returned DataFrame looks like the
    DF from report.get_remarks()'''
    if remarks.empty: return get_empty_remarks_df()
    oem = wea.get('oem')
    id = wea.get('id')
    year = wea.report.parent_project.get('year_id')
    kind = wea.report.inspection.get('kind')

    return remarks.loc[oem, id, year, kind]


def get_remarks_of_chapter(remarks, chapter):
    '''returns every remark in remarks that has the given chapter'''
    if remarks is None: return get_empty_remarks_df()
    corr_start = remarks.index.get_level_values('address').str.startswith(chapter)
    corr_len_of_start = (get_chapters_from_remarks(remarks, as_list=False, unique=False)
                         .str.len() == len(chapter))
    allowed_splits = chapter.count(db_split_char)+1 # +1 because title introduces extra split char
    no_subchapter = (remarks
                     .index                                
                     .get_level_values('address')
                     .str
                     .replace(db_split_char, '_SPLIT_')
                     .str
                     .count('_SPLIT_') <= allowed_splits)
    wanted_entries = corr_start & no_subchapter & corr_len_of_start
    return remarks.loc[wanted_entries]



def get_flagcolor(flags, only_PIEV=False):
    '''returns color corresponding to the worst flag in flags. priority: PEIV
    flags: list of flags'''
    
    colors = {'V': 'green2',
              'E': 'orange',
              'I': 'sky blue',
              'P': 'red',
              '3': 'yellow2',              # 0, 2, and 4 use non_setup, V and P entries
              'non_setup': 'gray54',
              'non_existant': 'snow',
              'others': 'light slate blue'}

    # priority: P, E, I, V
    if 'P' in flags or 'PP' in flags or 'PPP' in flags: return colors['P']
    if 'E' in flags: return colors['E']
    if 'I' in flags: return colors['I']
    if 'V' in flags: return colors['V']
    if only_PIEV: return None
    if '2' in flags: return colors['V']
    if '3' in flags: return colors['3']
    if '4' in flags: return colors['P']
    if '0' in flags: return colors['non_setup']
    return colors['others']



def get_overview(project) -> pd.DataFrame:
    rems = project.get_all_remarks(allowed_flags=['PPP', 'PP', 'P', 'I', 'V', 'E'],
                                   drop_cols=['image_names', 'author'])

    # concat flag and fulltext
    rems['remark'] = rems[['flag', 'fulltext']].agg(': '.join, axis=1)
    
    # get rid of create time
    rems = rems.droplevel('create_time').set_index(
        rems.groupby(['turbine_id', 'address']).cumcount().rename('ind'),
        append=True)
    rems = rems.droplevel(['oem', 'insp_year', 'inspection_type'])
    
    rems = rems['remark']
    overview = rems.unstack('turbine_id')

    # sort
    overview = sort_remarks_using_address(overview)

    # separate section an title
    address_particles = overview.index.get_level_values('address').str.split('|')
    overview['chapter'] = ['|'.join(molecule[:-1]) for molecule in address_particles]
    overview['title'] = [molecule[-1] for molecule in address_particles]

    overview = (overview
                .set_index(['chapter', 'title'], append=True)
                .droplevel(['address', 'ind'])
                .set_index(overview
                           .groupby(['chapter', 'title'])
                           .cumcount()
                           .rename('ind'),
                           append=True))
    return overview

def excel_overview(project):
    rems = get_overview(project)
    excel_path = f'{os.getcwd()}/overview {project.name}.xlsx'

    ids = rems.columns

    old_address_str = ''
    old_address = ()
    old_title = ''
    wb = pyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 30

    for i, id in enumerate(ids):
        letter = list('BCDEFGHIJKLMNOPQRSTUVWXYZ')[i]
        ws.column_dimensions[letter].width = 50
        cell = ws.cell(row=1, column=i+2)
        cell.value = id
        cell.font = pyxl.styles.Font(size=15, color='2A4879', bold=True)

    colors = {
        'P': pyxl.styles.PatternFill(start_color='FEFE54', end_color='FEFE54', fill_type='solid'),
        'PP': pyxl.styles.PatternFill(start_color='FD8A17', end_color='FD8A17', fill_type='solid'),
        'PPP': pyxl.styles.PatternFill(start_color='FE5050', end_color='FE5050', fill_type='solid'),
        'V': pyxl.styles.PatternFill(start_color='377E22', end_color='377E22', fill_type='solid'),
        'I': pyxl.styles.PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid'),
        'E': pyxl.styles.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')        
    }

    turb_id_cols = {id: j for j, id in enumerate(ids, 2)}

    i = 2
    for address_str, title, _ in rems.index:
        if address_str == old_address_str and title == old_title:
            continue
        if not address_str.startswith('Prüfbemerkungen|'):
            raise IndexError(f'address needs to be a Prüfbemerkung, not {address_str}')
        address = address_str[16:].split(db_split_char)
        for j, atom in enumerate(address):
            try: 
                if atom == old_address[j]: continue
                else: raise IndexError
            except IndexError:
                remaining_atoms = address[j:]
                for j, atom in enumerate(remaining_atoms, j):
                    cell = ws.cell(row=i, column=1)
                    cell.value = f'{(j+1)*' '}{atom}'
                    cell.font = pyxl.styles.Font(size=15-j,
                                                 color='2A4879',
                                                 bold=True)
                    i += 1
                break
        
        titlecell = ws.cell(row=i, column=1)
        titlecell.alignment = pyxl.styles.Alignment(vertical='top', wrap_text=True)
        titlecell.value = title 
        vals = rems.loc[address_str, title].stack()
        vals.name = 'text'
        vals_dict = multiindex2dict(vals
                                    .reset_index()
                                    .set_index(['text', 'turbine_id'])
                                    .index)
        
        used_rows = {id: [] for id in ids}

        for text, turb_ids in vals_dict.items():
            flag = text[:text.find(':')]
            try: turb_ids = (pd.Series(turb_ids)
                             .astype(int)
                             .sort_values()
                             .astype(str)
                             .to_list())
            except ValueError: pass
            row = i
            row_lists = {turb_id: used_rows[turb_id] for turb_id in turb_ids}
            while row in [row for row_list in row_lists.values() for row in row_list]:
                row += 1

            for j, turb_id in enumerate(turb_ids):
                if j == 0: celltext = gui_f.dbtext2displaytext(text)
                else: celltext = f's. {turb_ids[0]}'
                cell = ws.cell(row=row, column=turb_id_cols[turb_id])
                cell.value = celltext
                cell.fill = colors[flag]
                cell.alignment = pyxl.styles.Alignment(vertical='top',
                                                        horizontal='center',
                                                        wrap_text=True)
                used_rows[turb_id].append(row)
        i = max([row for rowlist in used_rows.values() for row in rowlist]) + 1

        old_address = address
        old_address_str = address_str
        old_title = title
    ws.freeze_panes = ws['A2']
    DEFAULT_FONT.name = 'Arial'
    wb.save(excel_path)


def get_overview_8p2_Inspect(project):
    rems = (project.get_all_remarks(allowed_flags=['PPP', 'PP', 'P', 'I', 'V', 'E'],
                                    drop_cols=['image_names', 'author'])
            .reset_index()
            .set_index(['fulltext', 'turbine_id', 'address'])
            .sort_index()
            .drop(['oem', 'insp_year', 'inspection_type', 'create_time', 'position'], axis=1)
            .unstack('turbine_id')
            .reorder_levels(['turbine_id', 0], axis=1)
            .sort_index(axis=1)
            .droplevel(1, axis=1))
    rems = sort_remarks_using_address(rems)
    return rems

def excel_overview_8p2_Inspect(project):
    rems = get_overview_8p2_Inspect(project)
    ids = rems.columns

    old_address = ()
    wb = pyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 50

    for i, id in enumerate(ids):
        letter = list('BCDEFGHIJKLMNOPQRSTUVWXYZ')[i]
        ws.column_dimensions[letter].width = 30
        cell = ws.cell(row=1, column=i+2)
        cell.value = id
        cell.font = pyxl.styles.Font(size=15, color='2A4879', bold=True)

    colors = {
        'P': pyxl.styles.PatternFill(start_color='FEFE54', end_color='FEFE54', fill_type='solid'),
        'PP': pyxl.styles.PatternFill(start_color='FD8A17', end_color='FD8A17', fill_type='solid'),
        'PPP': pyxl.styles.PatternFill(start_color='FE5050', end_color='FE5050', fill_type='solid'),
        'V': pyxl.styles.PatternFill(start_color='377E22', end_color='377E22', fill_type='solid'),
        'I': pyxl.styles.PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid'),
        'E': pyxl.styles.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')        
    }

    i = 2
    for fulltext, address_str in rems.index:
        if not address_str.startswith('Prüfbemerkungen|'):
            raise IndexError(f'address needs to be a Prüfbemerkung, not {address_str}')
        address = address_str[16:].split(db_split_char)[:-1]
        for j, atom in enumerate(address):
            try: 
                if atom == old_address[j]: continue
                else: raise IndexError
            except IndexError:
                remaining_atoms = address[j:]
                for j, atom in enumerate(remaining_atoms, j):
                    cell = ws.cell(row=i, column=1)
                    cell.value = f'{(j+1)*' '}{atom}'
                    cell.font = pyxl.styles.Font(size=15-j,
                                                 color='2A4879',
                                                 bold=True)
                    i += 1
                break
        
        textcell = ws.cell(row=i, column=1)
        textcell.alignment = pyxl.styles.Alignment(vertical='top', wrap_text=True)
        textcell.value = gui_f.dbtext2displaytext(fulltext) 
        vals = rems.loc[fulltext, address_str]
        
        for j, turb_id in enumerate(ids, 2):
            flag = vals[turb_id]
            if pd.isna(flag):
                continue
            flagcell = ws.cell(row=i, column=j)
            flagcell.value = flag
            flagcell.fill = colors[flag]
            flagcell.alignment = pyxl.styles.Alignment(vertical='center',
                                                       horizontal='center')


        i += 1
        old_address = address
    ws.freeze_panes = ws['A2']
    DEFAULT_FONT.name = 'Arial'
    wb.save(f'{os.getcwd()}/overview {project.name}_Insp.xlsx')


def get_default_remarks(checklist):
    default_on = checklist.default_state.dropna().index
    return checklist.loc[default_on]

def check_if_parts_in_db(wea) -> bool:
    db = load_parts()
    try:
        db.loc[idx[wea.oem, :, :, wea.id]]
        return True
    except KeyError:
        return False
        
def check_if_current_temperatures_in_db(wea, inspection):
    year = inspection.get_year()
    db = load_temperatures()
    slc = pd.DataFrame()
    try:
        slc = db.loc[idx[wea.oem, :, :, wea.id, year]]
    except KeyError:
        try:
            slc = db.loc[idx[wea.oem, :, :, wea.id, int(year)]]
        except KeyError: pass
    if slc.empty: return False
    return True
        

def get_all_inspectors_ever():
    '''return dict with all inspectors in inspections db.
    Key is role and Values is a list of persons having occupied that role'''
    inspections_db = load_inspections()
    inspectors = list(set([single_inspector[:2]
                            for inspectors_list in inspections_db.inspectors.values 
                                for single_inspector in eval(inspectors_list)]))
    inspectors_ind = pd.MultiIndex.from_tuples(inspectors).sort_values()
    return multiindex2dict(inspectors_ind)


def get_all_tower_types_ever(hybrid=False):
    '''returns a llist with all tower_types in turbine database.
    hybrid: bool, default False. If True, hybrid towers are included as a whole,
    else only their parts are included. (['Steel|Concrete', ...] vs. ['Steel', 'Concrete'])'''
    turbines_db = load_turbines()
    tower_types_raw = (turbines_db
                       .dropna(subset='tower_type')
                       .tower_type
                       .sort_values()
                       .unique())
    tower_types = []

    for tt in tower_types_raw:
        section_types = tt.split(db_split_char)
        section_types = [st.strip() for st in section_types]
        if hybrid:
            tt = db_split_char.join(section_types)
            if tt not in tower_types: tower_types.append(tt)
            continue
        for section_type in section_types:
            if section_type not in tower_types: tower_types.append(section_type)
    
    return tower_types

def get_all_oem_model_ever():
    turbines_db = load_turbines().reset_index()
    return (turbines_db
            .groupby(['oem', 'model'])
            .head(1)[['oem', 'model']]
            .dropna()
            .agg(' '.join, axis=1)
            .to_list())

def get_all_scopes_ever():
    inspections_db = load_inspections()
    scopes = list(inspections_db.scope.dropna().unique())
    return scopes

def get_all(what, **kwargs):
    if what == 'tower_type':
        return get_all_tower_types_ever(**kwargs)
    if what == 'inspector':
        return get_all_inspectors_ever()
    if what == 'oemmodel':
        return get_all_oem_model_ever()
    if what == 'scope':
        return get_all_scopes_ever()
    db = load_turbines().reset_index()
    return list(db[what].dropna().unique())

def get_all_db_values() -> list:
    '''get list of all checklist entries and all remark_db entries
    format:
        +++ occurrences of text +++ text'''
    cl = load_checklist()
    rems = load_remarks()
    turbs = load_turbines()
    all_entries = {}
    used_fulltexts = []
    db_values = []
    for ind in cl.index:
        entry = cl.loc[ind]
        fulltext = entry.fulltext
        address = f'{ind[0].replace(db_split_char, ' > ')} > {ind[1]}'        
        
        key = f'+++ {address} +++\n{fulltext}'.strip()
        if key.lower() in used_fulltexts: continue

        all_entries[key] = ['Vorlage']
        used_fulltexts.append(key.lower())

    for ind in rems.index:
        entry = rems.loc[ind]
        fulltext = entry.fulltext
        address = ind[4].replace(db_split_char, ' > ')

        key = f'+++ {address} +++\n{fulltext}'.strip()
        used_in = all_entries.get(key, [])

        oem = ind[0]
        try: turb_model = turbs.loc[oem, ind[1]]['model'].iloc[0]
        except KeyError: turb_model = '(Modell unbekannt)'
        turb = f'{oem} {turb_model}'
        if turb in used_in: continue

        used_in.append(turb)
        all_entries[key] = used_in

    
    for key, used_in in all_entries.items():
        in_checklist = True
        try: used_in.remove('Vorlage')
        except ValueError: in_checklist = False
        used_in = sorted(used_in)
        if in_checklist: used_in.insert(0, 'Vorlage')
        used_in_str = f'\n(Verwendet in: {', '.join(used_in)})'
        db_values.append(f'{key}{used_in_str}')
        
    return db_values



def get_oem_model_X_dict(oem, what):
    turbs = load_turbines()
    try: options = (turbs.loc[[oem]]
                    .groupby(['model', what])
                    .head(1)
                    .set_index('model')[what]
                    .sort_index())
    except KeyError: options = pd.Series()

    return {model: options.loc[[model]].sort_values().to_list()\
                for model in options.index}



def get_oem_model_dict():
    '''return a dictionary with oems an dcooresponding models as present in
    turbines_db. dict is sorted according to frequency of oem model combination'''
    ind_available_models_in_wea_db = (load_turbines()
                                      .groupby(['oem', 'model'])
                                      .count()
                                      .sort_values('location',
                                                   ascending=False)
                                      .index)
    return multiindex2dict(ind_available_models_in_wea_db)

def get_ids_of_model(oem, model):
    '''get a list of serial numbers for given oem and model'''
    t = load_turbines()
    try: t_oem = t.loc[oem]
    except KeyError: return []
    ids = t_oem.loc[t_oem.model == model].index
    return ids.to_list()



def sort_dict(_dict, sorted_list) -> dict:
    '''sort a dictionary according to the order given by sorted_list. Put unknown dict keys at the end.'''
    sorted_dict = {}
    dict_keys = _dict.keys()
    for entry in sorted_list:
        if entry in dict_keys: sorted_dict[entry] = _dict[entry]
    for key in dict_keys:
        if key not in sorted_dict:
            sorted_dict[key] = _dict[key]
    return sorted_dict
    


def compare_temperatures(wea_type: str, ax=None,
                         highlight_ids=[],
                         highlight_old=False,
                         all_subtypes=True,
                         oem=None):
    temperatures_db = load_temperatures()
    if not ax:
        fig, ax = plt.subplots()    
    if isinstance(highlight_ids, int):
        highlight_ids = [str(highlight_ids)]
    elif isinstance(highlight_ids, str):
        highlight_ids = [highlight_ids]

    if all_subtypes:
        wea_type = (temperatures_db
                    .index
                    .unique('model')[temperatures_db
                                     .index
                                     .unique('model')
                                     .str
                                     .contains(wea_type)])
        
    if oem: ind = idx[oem, wea_type]
    else: ind = idx[:, wea_type]

    df = temperatures_db.loc[ind, :]

    if highlight_old:
        to_boxplot = (df.drop(highlight_ids, level='turbine_id'))
        to_highlight = (df
                        .loc[idx[:, wea_type, :, highlight_ids], :]
                        .reset_index())
    else:
        _to_highlight = []
        to_boxplot = df
        for turbine_id in highlight_ids:
            most_recent_year = (df
                                .loc[idx[:, :, :, turbine_id], :]
                                .groupby('insp_year')
                                .head(1)
                                .droplevel(['oem', 'model', 'name', 'turbine_id'])
                                .sort_index(level='insp_year', ascending=False)
                                .index[0])
            _highlight = df.loc[idx[:, :, :, turbine_id, most_recent_year], :]
            _to_highlight.append(_highlight)
            to_boxplot = to_boxplot.drop(_highlight.index)
        if len(highlight_ids) > 0:
            to_highlight = pd.concat(_to_highlight).reset_index()

    temp_order = get_order('temperatures')
    ind = [t for t in temp_order if t in to_boxplot.index.unique('name')]
    to_boxplot = (to_boxplot
                  .reorder_levels(['name', 'oem', 'model', 'turbine_id', 'insp_year'])
                  .loc[ind]
                  .reset_index())

    ax.tick_params(axis='x', labelrotation=90)
    ax = sns.boxenplot(to_boxplot,
                       x='name', y='value',
                       fill=False,
                       hue=.75,
                       palette='dark:black',
                       linewidth=1,
                       zorder=3,
                       legend=True if highlight_ids else False,
                       ax=ax)
    ax.yaxis.grid(True, zorder=1)   # show the horizontal gridlines
    ax.xaxis.grid(False)            # hide the vertical gridlines
    
    if len(highlight_ids) == 0:
        return ax
    
    ax = sns.scatterplot(to_highlight,
                         x='name', y='value',
                         hue='turbine_id',
                         style='insp_year' if highlight_old else None,
                         s=50,
                         zorder=4,
                         ax=ax,
                         )
    # adjust legend
    h, l = ax.get_legend_handles_labels()
    l[0] = wea_type
    ax.legend(h[1:], l[1:])
    return ax




def turbinedata2inspectiondb(columnname_from_turb_db, isnumeric=False, insp=None):
    if insp is None: # insp cna be given, eg if multiple additional columns are being added sequencially
        insp = load_inspections()
    turbs = load_turbines()

    insp[columnname_from_turb_db] = [np.nan]*len(insp)

    def set_data(ind):
        _, _, oem, sn = ind
        try: value = turbs.loc[idx[oem, sn], columnname_from_turb_db].iloc[0]
        except KeyError: return
        if isnumeric: value = float(value)
        else: value = str(value)
        insp.loc[ind, columnname_from_turb_db] = value

    insp.apply(lambda row: set_data(row.name), axis=1)
    return insp

def coordinates2inspectiondb(insp=None):
    if insp is None:
        insp = load_inspections()
    turbs = load_turbines()
    insp['lat'] = [np.nan]*len(insp)
    insp['long'] = [np.nan]*len(insp)

    def set_coordinates(ind):
        _, _, oem, sn = ind
        try: lat, long = eval(turbs.loc[idx[oem, sn], 'coordinates'].iloc[0])
        except KeyError: return
        except TypeError: return # happens is coords are nan
        insp.loc[ind, 'long'] = long
        insp.loc[ind, 'lat'] = lat

    insp.apply(lambda row: set_coordinates(row.name), axis=1)
    return insp


def rotorarea2inspectiondb(insp=None):
    insp = turbinedata2inspectiondb('rotor_diam', isnumeric=True, insp=insp)
    insp['rotor_area'] = insp['rotor_diam']**2 * np.pi/4
    return insp


def powerfac2inspdata(insp=None):
    if insp is not None:
        if not 'rated_power' in insp.columns:
            insp = turbinedata2inspectiondb('rated_power', isnumeric=True, insp=insp)
    else:   insp = turbinedata2inspectiondb('rated_power', isnumeric=True)
    insp = turbinedata2inspectiondb('model', insp=insp)
    insp = rotorarea2inspectiondb(insp=insp)
    insp['powerfac'] = insp['wea_output']/insp['wea_hours']/insp['rated_power']
    insp['power_per_area'] = insp['wea_output']/insp['wea_hours']/insp['rotor_area']
    return insp

def startupyear2inspdata(insp=None):
    if insp is not None:
        insp = turbinedata2inspectiondb('startup_date', insp=insp)
    else: insp = turbinedata2inspectiondb('startup_date')
    insp['startup_year'] = pd.DatetimeIndex(
                        pd.to_datetime(insp['startup_date'],
                        format='mixed', dayfirst=True, errors='coerce')).year
    insp.drop('startup_date', axis='columns', inplace=True)

    return insp

def inspyear2inspdata(insp=None):
    if insp is None:
        insp = load_inspections()
    insp['inspection_year'] = pd.DatetimeIndex(
        pd.to_datetime(insp.index.get_level_values('insp_year'), format='mixed',
                       dayfirst=True, errors='coerce')).year
    return insp


def monthyear2datetime(monthyear_str):
    '''converts strings like '01/24' to a pandas datetime object (01.01.2024)
    handles mm/yy or mm/yyyy'''
    _, year = monthyear_str.split('/')
    if len(year) == 2: year_format = 'y'
    elif len(year) == 4: year_format = 'Y'
    else: raise ValueError(f'monthyear falsch formatiert. Erwartet mm/yy oder mm/yyyy, nicht {monthyear_str}')
    return pd.to_datetime(monthyear_str, format=f'%m/%{year_format}')

def add_months(og_timestamp, months_to_add):
    return og_timestamp+pd.Timedelta(365.25/12, 'days')*months_to_add

def get_timeline_figure(timeline_data: dict, start: str, end: str):
    mpl.rcParams['font.sans-serif'] = 'Arial'
    mpl.rcParams['font.family'] = 'sans-serif'
    for fontsize in ['font.size', 'axes.titlesize',
                     'axes.labelsize', 'xtick.labelsize']:
        plt.rcParams.update({fontsize: 11})


    keys = list(timeline_data.keys())
    nrows = len(keys)

    # Convert start and end dates to datetime objects
    start_date = datetime.strptime(start, '%m/%y')
    end_date = datetime.strptime(end, '%m/%y')

    # Create figure and axis
    fig, ax = plt.subplots(figsize=(5.5, .5+nrows/3), layout='constrained')    
    # format axis
    ax.spines[['top', 'left', 'right', 'bottom']].set_visible(False)
    # Assign y positions to each key
    y_positions = list(range(nrows))
    ax.set_yticks(y_positions)
    ax.set_ylim([-nrows/10, y_positions[-1]+nrows/10])
    ax.set_yticklabels([key.replace('|', '\n') for key in keys])
    ax.invert_yaxis()  # Top key first
    ax.tick_params(axis='y', which='both', length=0)

    # Set x-axis limits and format
    datespan = (end_date-start_date).days / 30 # months
    ax.set_xlim(start_date-relativedelta(months=1), end_date+relativedelta(months=round(datespan/20)))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
    ax.xaxis.set_major_locator(mdates.AutoDateLocator())
    ax.tick_params(axis='x', which='both', direction='inout', length=6, labelrotation=45)

    # Custom x-ticks handling    
    # Get potential ticks then filter
    raw_ticks = mdates.AutoDateLocator().tick_values(start_date, end_date)
    min_distance = round(datespan*1.5)  # Minimum days between edge ticks and first/last interior ticks
    
    # Convert to datetime and filter
    filtered_ticks = []
    for tick in raw_ticks:
        tick_date = mdates.num2date(tick).replace(tzinfo=None)
        if tick_date < start_date or tick_date > end_date:
            continue
            
        # Check proximity to edges
        days_from_start = (tick_date - start_date).days
        days_from_end = (end_date - tick_date).days
        
        # Always keep start/end dates
        if days_from_start == 0 or days_from_end == 0:
            filtered_ticks.append(tick_date)
            continue
            
        # Filter nearby ticks
        if days_from_start > min_distance and days_from_end > min_distance:
            filtered_ticks.append(tick_date)
    
    # Ensure start/end are always included
    filtered_dates = [start_date, end_date] + [
        t for t in filtered_ticks 
        if t not in {start_date, end_date}
    ]
    filtered_dates.sort()

    ax.xaxis.set_major_locator(FixedLocator(mdates.date2num(filtered_dates)))
    ax.tick_params(axis='x', which='both', direction='inout', length=6, labelrotation=45)


    # Plot each entry
    for y_pos, key in zip(y_positions, keys):
        entry = timeline_data[key]
        interval = None
        date_strs = entry
        
        # Check if the first element is an interval (integer)
        if isinstance(entry[0], int):
            interval = entry[0]
            date_strs = entry[1:]
        
        # Convert date strings to datetime objects
        dates_list = [datetime.strptime(ds, '%m/%y') for ds in date_strs]
        
        # Plot vertical markers
        for date in dates_list:
            ax.plot(date, y_pos, 'k|', markersize=8, markeredgewidth=2, zorder=3)
        
        # Plot blue interval lines
        if interval is not None:
            for date in dates_list:
                end_date_line = min([date + relativedelta(months=+interval),
                                     end_date])
                ax.hlines(
                    y=y_pos, xmin=date, xmax=end_date_line,
                    colors='tab:blue', linewidth=2, alpha=0.35, zorder=2)
    for date in [start_date, end_date]:
        ax.axvline(date, c='tab:gray', ls='--', lw=1, zorder=1, alpha=.5)


    # add arrow as x axis
    xmin, xmax = ax.get_xlim()
    ymin, _ = ax.get_ylim()
    arrow = mpatches.FancyArrowPatch(
        (xmin, ymin), (xmax, ymin),
        arrowstyle='-|>',
        mutation_scale=10,
        color='black',
        linewidth=0.8,
        clip_on=False
    )
    ax.add_patch(arrow)
    return fig

    
def get_curr_time():
    return pd.to_datetime(time.time(), unit='s')

def init_databases():
    path = f'{mainpath}/databases'
    def init_data_file(name, data):
        """Initialize a file if it doesn't exist."""
        if f'{name}.csv' not in os.listdir(path):
            save_db(data, name)
            print(f'init: created new {name} database.')
        else: print(f'init: present: {name} database.')

    def init_order_file(order_name, order_data):
        """Initialize order file if it doesn't exist."""
        if f'order_{order_name}.txt' not in os.listdir(path):
            with open(f'{path}/order_{order_name}.txt', 'w', encoding='utf-8') as f:
                f.write(str(order_data).replace('], \'', '],\n\'').replace("', '", "',\n'"))
            print(f'init: created new {order_name} order.')
        else:
            print(f'init: present: {order_name} order.')

    databases = {
        "checklist": {
            "file_name": "checklist.csv",
            "data": pd.DataFrame().from_dict(get_default_checklist_dict(), orient='tight'),
            "order_name": "chapters",
            "order_data": get_default_chapter_order()
        },
        "remarks": {
            "file_name": "remarks.csv",
            "data": pd.DataFrame().from_dict(get_default_remarks_dict(get_curr_time()),
                                             orient='tight'),
            "order_name": None,
            "order_data": None
        },
        "components": {
            "file_name": "components.csv",
            "data": pd.DataFrame().from_dict(get_default_components_dict(), orient='tight'),
            "order_name": "components",
            "order_data": get_default_component_order()
        },
        "temperatures": {
            "file_name": "temperatures.csv",
            "data": pd.DataFrame().from_dict(get_default_temperatures_dict(), orient='tight'),
            "order_name": "temperatures",
            "order_data": get_default_temperature_order()
        },
        "turbines": {
            "file_name": "turbines.csv",
            "data": pd.DataFrame().from_dict(get_default_turbines_dict(), orient='tight'),
            "order_name": None,
            "order_data": None
        },
        "inspections": {
            "data": pd.DataFrame().from_dict(get_default_inspections_dict(), orient='tight'),
            "order_name": None,
            "order_data": None
        }
    }  

    path = f'{mainpath}/databases'
    for name, details in databases.items():
        init_data_file(name, details["data"])
        if details["order_name"] and details["order_data"]:
            init_order_file(details["order_name"], details["order_data"])


def init_config():
    if 'config.py' in os.listdir(f'{mainpath}'):
        print('init: present: config.py')
        return
    config_str = get_default_config()
    with open(f'{mainpath}/config.py', 'w', encoding='utf-8') as f:
        f.write(config_str)


def get_flag_stripplot(flag, oem, model, insp_type, curr_id=None,
                        all_remarks=None):
    '''get a violinplot showing the distribution of curr flag. Comparison only
    between same wea and same inspection type.
    flag (str): P, I or E
    oem (str): turbine oem
    model (str): turbine model
    inspection_type (str)
    curr_id (str): default None, place marker for current wea
    all_remarks (pd.DataFrame): remark DataFrame as returned by load_remarks()
    returns pyplot figure
    '''

    ids = get_ids_of_model(oem, model)
    rems = get_remarks_from_ids(oem, ids, remarks=all_remarks)
    rems = rems.loc[rems.index.get_level_values('inspection_type') == insp_type]

    flag_counts = (rems[rems.flag==flag]
                   .flag
                   .groupby(['oem', 'turbine_id', 'inspection_type'])
                   .value_counts())
    
    marker_pos = None
    if curr_id:
        try:
            marker_pos = flag_counts.loc[oem, curr_id].values[0]
            flag_counts = flag_counts.drop(curr_id, level='turbine_id')
        except: pass

    fig = plt.figure()
    fig.set_size_inches(1, .25)
    ax = plt.Axes(fig, [0., 0., 1., 1.])
    ax.set_axis_off()
    fig.add_axes(ax)

    sns.stripplot(flag_counts.reset_index(), x='count', orient='horizontal',
                   ax=ax, alpha=.5)
    
    if marker_pos: ax.axvline(marker_pos, lw=2, c='tab:orange', zorder=4)

    og_xlow, og_xhigh = ax.get_xlim()
    diff = og_xhigh - og_xlow
    new_xlim = (og_xlow-.1*diff, og_xhigh+.1*diff)
    ax.set(xlim=new_xlim)
    return fig




# f = get_flag_stripplot('P', 'Enercon', 'E82', 'ZOP')
# f.savefig('C:/Users/tadej/Documents/example_counter.png', dpi=300)





    




    
